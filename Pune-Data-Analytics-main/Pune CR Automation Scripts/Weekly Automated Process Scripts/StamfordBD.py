#!/usr/bin/env python
# coding: utf-8

# In[1]:


import snowflake.connector
import pandas as pd
# Gets the version
conn = snowflake.connector.connect(
    user="vinayak.laxmeshwar@zocdoc.com",
    password="do not enter your password. it will authenticate with the browser, however this parameter cannot be an empty string.",
    account="zocdoc_001.us-east-1",
    authenticator="externalbrowser",
    database="cistern"
    )


# In[2]:


from datetime import datetime, timedelta
day = datetime.today().strftime('%Y-%m-%d')
dt = datetime.strptime(day, '%Y-%m-%d')
end = dt - timedelta(days=dt.weekday())
start = end - timedelta(days=7)
print(start.strftime('%Y-%m-%d'))
print(end.strftime('%Y-%m-%d'))


# In[3]:


try:
    sql=f"""SELECT
       distinct a.appointment_id,
        a.monolith_parent_strategic_id,
        a.parent_strategic_name,
        a.provider_npi as Provider_NPI
         ,monolith_professional_id
         , a.provider_first_name as Provider_First_Name
         , a.provider_last_name as Provider_Last_Name
         , a.booking_specialty_name as Specialty
         , a.procedure_name as Procedure_Name
         , a.insurance_carrier_name as Insurance_Carrier
         , a.insurance_plan_name as Insurance_Plan
         , a.insurance_plan_type as Insurance_Type
         , CASE WHEN a.is_new_to_provider_from_pe_vw = TRUE then 'New' else 'Existing' END as New_Or_Existing_Patient
         , a.appointment_created_timestamp_local as Booking_Time
         , CASE WHEN a.is_appointment_created_time_after_hours_local = TRUE then 'After Hours' else 'Business Hours' END as After_Hours_vs_Business_Hours_Booking
         , a.platform as Device_Type
         , a.referrer_type_category as Booking_Source
         , CASE
                     WHEN a.is_non_chargeable_cancellation_from_pe_vw = TRUE then 'Non_Chargable_Patient_Cancellation'
                     WHEN a.cancellation_reason like 'Patient_%' then 'Cancelled_by_Patient'
                     WHEN a.cancellation_reason like 'Provider_%' then 'Cancelled_by_Provider'
                     WHEN a.appointment_outcome = 'RealizedAppointment' then 'Confirmed'
                     WHEN a.appointment_outcome = 'BookingFailed' then 'Rescheduling Error'
                     WHEN a.appointment_outcome is null then 'Upcoming Appointment'
                         else a.appointment_outcome end as Booking_Outcome
         , CASE
                     WHEN a.cancellation_reason like 'Patient_%' then 'Patient'
                     WHEN a.cancellation_reason like 'Provider_%' then 'Provider'
                     else null end as Cancellation_Initiator
         , a.cancellation_reason as Cancellation_Reason
         , (a.hours_between_creation_time_and_initial_appointment_time / 24.00) as Booking_Lead_Time_in_Days
         , a.latest_appointment_time_local as Appointment_Time
         , a.latest_event_timestamp_utc as Latest_Appointment_Change_Time
                     ,a.APPOINTMENT_OUTCOME_TIMESTAMP_UTC
                     ,convert_timezone('America/New_York', a.APPOINTMENT_OUTCOME_TIMESTAMP_UTC) as APPOINTMENT_OUTCOME_TIMESTAMP_EASTERN
         , a.realized_cost as Booking_Cost -- is this the correct field?
         , CASE WHEN a.is_premium_booking = TRUE then 'Premium Booking' else null END as Premium_Booking
         , CASE WHEN a.is_spo_booking = TRUE then 'SPO Booking' else null END as SPO_Booking
         , CASE WHEN a.is_virtual_location = TRUE then 'Video Visit' else 'In Person Visit' end as Visit_Type
         , a.practice_name as Practice_Name
         ,a.MONOLITH_PROVIDER_LOCATION_ID
         , a.provider_location_address as Street_Address
         , a.provider_location_city as City
         , z.county as County
         , a.provider_location_state as State
         , a.provider_location_zip_code as Zip_Code
         , a.strategic_name as Client_Name
         , a.session_id  as Session_ID
         , CASE WHEN a.referrer_type_category = 'Marketplace' then null else s.white_label_directory_id end as White_Label_ID
         , CASE WHEN a.referrer_type_category = 'Marketplace' then null else s.landing_page_url end as Landing_Page_URL
         , CASE WHEN a.referrer_type_category = 'Marketplace' then null else s.landing_page_url_network_location end as Landing_Page_URL_Network_Location
         , CASE WHEN a.referrer_type_category = 'Marketplace' then null else s.landing_page_url_path_component_1 end as Landing_Page_URL_Component_1
         , CASE WHEN a.referrer_type_category = 'Marketplace' then null else s.landing_page_url_path_component_2 end as Landing_Page_URL_Component_2
         , CASE WHEN a.referrer_type_category = 'Marketplace' then null else s.referrer_url end as Referrer_URL
         , CASE WHEN a.referrer_type_category = 'Marketplace' then null else s.platform end as Platform_Type
         , CASE WHEN a.referrer_type_category = 'Marketplace' then null else s.platform_detail end as Platform_Category
         , CASE WHEN a.referrer_type_category = 'Marketplace' then null else s.session_start_timestamp_utc end as Session_Start_Timestamp
         , CASE WHEN a.referrer_type_category = 'Marketplace' then null else s.tracking_id end as Tracking_ID
         , CASE WHEN a.referrer_type_category = 'Marketplace' then null else s.utm_campaign end as UTM_Campaign
         , CASE WHEN a.referrer_type_category = 'Marketplace' then null else s.utm_medium end as UTM_Medium
         , CASE WHEN a.referrer_type_category = 'Marketplace' then null else s.utm_source end as UTM_Source
         --, a.strategic_id
         --, a.monolith_strategic_id
         --, a.parent_strategic_id
         --, a.monolith_parent_strategic_id
     FROM APPOINTMENT.appointment_summary_commercial_vw as a
     LEFT JOIN user_behavior.session as s
     ON a.session_id = s.session_id
     LEFT JOIN public.zip_cbsa_mapping as z
        ON a.provider_location_zip_code = z.zip
     WHERE a.is_created_appointment = 'TRUE'
     AND a.appointment_created_timestamp_local between '{start.strftime('%Y-%m-%d')}' and '{end.strftime('%Y-%m-%d')}'
     AND procedure_id NOT IN ('pc_3BvV0dlIbkC_Hj_HQ8V4nB', 'pc_m459Tgq50kGv-faPIAb_3h','pc_wloLfj5vbUmMiwvX709pXx','pc_m0Wv6TPsc0iu-5Ju3vj7wh')
     --If the parent ID is filled out, use this line
     --AND a.parent_strategic_id_with_current_mapping in ('st_wFW-WpkaVUGExUnnHJBbJx')
     --If the regular ID is filled out or if you need to add multiple IDs, use this line
     AND a.monolith_strategic_id_with_current_mapping IN ('128')
     order by Booking_Time;"""
    
except:
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email import encoders
    import os.path
    email = 'vinayak.laxmeshwar@zocdoc.com'
    password = 'rubzzqumlkcrzagk'
    send_to_email = "vinayak.laxmeshwar@zocdoc.com, soumya.singh@zocdoc.com"
    subject = 'Stamford booking download'
    message = 'The code failed to run successfully'
    
    msg = MIMEMultipart()
    msg['From'] = email
    msg['To'] = send_to_email
    msg['Subject'] = subject

    msg.attach(MIMEText(message, 'plain'))
    
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(email, password)
    text = msg.as_string()
    server.sendmail(email, send_to_email.split(","), text)
    server.quit()


# In[4]:


cur = conn.cursor().execute(sql).fetchall()
df = pd.read_sql(sql, conn)


# In[5]:


df


# In[6]:


df.head(10)


# In[ ]:


df['APPOINTMENT_OUTCOME_TIMESTAMP_EASTERN'] = pd.to_datetime(df['APPOINTMENT_OUTCOME_TIMESTAMP_EASTERN'], utc=True)


# In[ ]:



df['APPOINTMENT_OUTCOME_TIMESTAMP_EASTERN'] = pd.to_datetime(df.APPOINTMENT_OUTCOME_TIMESTAMP_EASTERN, format='%Y-%m-%d %H:%M:%S')


# In[ ]:


from datetime import datetime, timedelta
import pytz

df['APPOINTMENT_OUTCOME_TIMESTAMP_EASTERN'] = df['APPOINTMENT_OUTCOME_TIMESTAMP_EASTERN'].dt.tz_convert('America/New_York')


# In[ ]:


##df['APPOINTMENT_OUTCOME_TIMESTAMP_EASTERN'] = pd.Series(df['APPOINTMENT_OUTCOME_TIMESTAMP_EASTERN']).dt.round("S") 


# In[ ]:


from datetime import datetime, timedelta
day = datetime.today().strftime('%Y-%m-%d')
dt = datetime.strptime(day, '%Y-%m-%d')
end_date = dt - timedelta(days=dt.weekday())
end_date = end_date - timedelta(days=1)
start_date = end_date - timedelta(days=6)
start_date= str(start_date.strftime('%m.%d.%Y'))
end_date= str(end_date.strftime('%m.%d.%Y'))


# In[ ]:


print(start_date)
print(end_date)


# In[ ]:


date_columns = df.select_dtypes(include=['datetime64[ns, UTC]']).columns
for date_column in date_columns:
    df[date_column] = df[date_column].apply(str)


# In[ ]:


df.replace({'NaT': ' '}, inplace=True)
df.to_excel(r'G:\My Drive\Stamford BD\Stamford Bookings Download ' + start_date + ' to ' + end_date+ '.xlsx', index= False )


# In[ ]:


import openpyxl
from openpyxl.styles import Font, Color, PatternFill,Alignment 

wb = openpyxl.load_workbook(r'G:\My Drive\Stamford BD\Stamford Bookings Download ' + start_date + ' to ' + end_date+ '.xlsx')
ws = wb['Sheet1']
redFill = PatternFill(start_color='e6f3ff',
                   end_color='e6f3ff',
                   fill_type='solid')


# Enumerate the cells in the second row
for cell in ws["1:1"]:
    cell.fill = redFill
    
wb.save(r'G:\My Drive\Stamford BD\Stamford Bookings Download ' + start_date + ' to ' + end_date+ '.xlsx')


# In[ ]:


import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os.path

email = 'vinayak.laxmeshwar@zocdoc.com'
password = 'rubzzqumlkcrzagk'
to = "andrew.card@zocdoc.com,selena.christian@zocdoc.com"
cc = "tanvi.malik@zocdoc.com,enterprisesupport@zocdoc.com"
subject = 'Stamford Bookings Download- ' + start_date + ' to ' + end_date+'.'
message = """Hi all,

Please find the attached bookings download data for Stamford.

Feel free to reach out if you have any questions.

Thanks & Regards,
Vinayak R Laxmeshwar"""

file_location = r'G:\My Drive\Stamford BD\Stamford Bookings Download ' + start_date + ' to ' + end_date+ '.xlsx'

rcpt = cc.split(",") + [to]

msg = MIMEMultipart()
msg['From'] = email
msg['To'] = to
msg['Cc'] = cc
msg['Subject'] = subject

msg.attach(MIMEText(message, 'plain'))

# Setup the attachment
filename = os.path.basename(file_location)
attachment = open(file_location, "rb")
part = MIMEBase('application', 'octet-stream')
part.set_payload(attachment.read())
encoders.encode_base64(part)
part.add_header('Content-Disposition', "attachment; filename= %s" % filename)

# Attach the attachment to the MIMEMultipart object
msg.attach(part)

server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login(email, password)
text = msg.as_string()
server.sendmail(email, rcpt, text)
server.quit()


# In[ ]:




