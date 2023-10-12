#!/usr/bin/env python
# coding: utf-8

# In[2]:


import pandas as pd
from pandas import DataFrame
import snowflake.connector
import sqlite3

# Gets the version
conn = snowflake.connector.connect(
    user="vinayak.laxmeshwar@zocdoc.com",
    password="do not enter your password. it will authenticate with the browser, however this parameter cannot be an empty string.",
    account="zocdoc_001.us-east-1",
    authenticator="externalbrowser",
    database="cistern"
    )

con = sqlite3.connect('rhcp.db')


# In[3]:


from datetime import datetime, timedelta
day = datetime.today().strftime('%Y-%m-%d')
dt = datetime.strptime(day, '%Y-%m-%d')
end = dt - timedelta(days=dt.weekday())
start = end - timedelta(days=7)
print(start.strftime('%Y-%m-%d'))
print(end.strftime('%Y-%m-%d'))


# In[4]:


df=pd.read_sql(f"""SELECT distinct a.appointment_id
    , a.provider_npi as Provider_NPI
 ,monolith_professional_id
    , a.provider_first_name as Provider_First_Name
    , a.provider_last_name as Provider_Last_Name
    ,provider_first_name || ' ' || provider_last_name as "Provider_Full_Name"
    , a.booking_specialty_name as Specialty
    , a.procedure_name as Procedure_Name
    , a.insurance_carrier_name as Insurance_Carrier
    , a.insurance_plan_name as Insurance_Plan
    , a.insurance_plan_type as Insurance_Type
    , CASE WHEN a.is_new_to_provider_from_pe_vw = TRUE then 'New' else 'Existing' END as New_Or_Existing_Patient
    , a.appointment_created_timestamp_local as Booking_Time
    , CASE WHEN a.is_appointment_created_time_after_hours_local = TRUE then 'After Hours' else 'Business Hours' END as After_Hours_vs_Business_Hours_Booking
    , a.platform as Device_Type
    , a.referrer_type_category as Booking_Source
    , case 
                when a.referrer_type_category = 'Marketplace' then '1'else '0'end as Marketplace_Bookings
   , case 
                when a.referrer_type_category = 'White Label' then '1'else '0'end as White_Label_Bookings
    , case 
                when a.referrer_type_category = 'Widget' then '1'else '0'end as Widget_Bookings
    , CASE
                WHEN a.is_non_chargeable_cancellation_from_pe_vw = TRUE then 'Non_Chargable_Patient_Cancellation'
                WHEN a.cancellation_reason like 'Patient_%' then 'Cancelled_by_Patient'
                WHEN a.cancellation_reason like 'Provider_%' then 'Cancelled_by_Provider'
                WHEN a.appointment_outcome = 'RealizedAppointment' then 'Confirmed'
                WHEN a.appointment_outcome = 'BookingFailed' then 'Rescheduling Error'
                WHEN a.appointment_outcome is null then 'Upcoming Appointment'
                    else a.appointment_outcome end as Booking_Outcome
    , CASE
                WHEN a.cancellation_reason like 'Patient_%' then 'Patient'
                WHEN a.cancellation_reason like 'Provider_%' then 'Provider'
                else null end as Cancellation_Initiator
    , a.cancellation_reason as Cancellation_Reason
     , a.latest_appointment_time_local as Appointment_Time
    , a.latest_event_timestamp_utc as Latest_Appointment_Change_Time
                 ,a.APPOINTMENT_OUTCOME_TIMESTAMP_UTC
                 ,convert_timezone('America/New_York', a.APPOINTMENT_OUTCOME_TIMESTAMP_UTC) as APPOINTMENT_OUTCOME_TIMESTAMP_EASTERN
    , a.realized_cost as Booking_Cost
    , CASE WHEN a.is_premium_booking = TRUE then 'Premium Booking' else null END as Premium_Booking
    , CASE WHEN a.is_virtual_location = TRUE then 'Video Visit' else 'In Person Visit' end as Visit_Type
    , a.practice_name as Practice_Name
    , a.provider_location_state as State
     , a.strategic_name as Client_Name
    , a.monolith_parent_strategic_id
    
FROM appointment.appointment_summary_commercial_vw as a
LEFT JOIN user_behavior.session as s
ON a.session_id = s.session_id
LEFT JOIN public.zip_geography as z
ON a.provider_location_zip_code = z.zip_code
left JOIN PROVIDER_ANALYTICS.practice_parent_entity_mapping_vw as e
ON a.monolith_provider_id = e.MONOLITH_PROVIDER_ID
left join provider.specialty_category sc on a.booking_specialty_id = sc.specialty_id
WHERE a.is_created_appointment = 'TRUE'
AND a.procedure_id NOT IN ('pc_3BvV0dlIbkC_Hj_HQ8V4nB', 'pc_m459Tgq50kGv-faPIAb_3h','pc_wloLfj5vbUmMiwvX709pXx','pc_m0Wv6TPsc0iu-5Ju3vj7wh')
AND a.appointment_created_timestamp_local between '{start.strftime('%Y-%m-%d')}' and '{end.strftime('%Y-%m-%d')}'
AND e.SALESFORCE_ULTIMATE_PARENT_ACCOUNT_ID like '%0013200001JDdxNAAT%'

Order by booking_time;""",conn)


# In[6]:


df['APPOINTMENT_OUTCOME_TIMESTAMP_EASTERN'] = pd.to_datetime(df['APPOINTMENT_OUTCOME_TIMESTAMP_EASTERN'], utc=True)


# In[7]:


df['APPOINTMENT_OUTCOME_TIMESTAMP_EASTERN'] = pd.to_datetime(df.APPOINTMENT_OUTCOME_TIMESTAMP_EASTERN, format='%Y-%m-%d %H:%M:%S')


# In[8]:


from datetime import datetime, timedelta
import pytz

df['APPOINTMENT_OUTCOME_TIMESTAMP_EASTERN'] = df['APPOINTMENT_OUTCOME_TIMESTAMP_EASTERN'].dt.tz_convert('America/New_York')


# In[9]:


df['APPOINTMENT_OUTCOME_TIMESTAMP_EASTERN'] = pd.Series(df['APPOINTMENT_OUTCOME_TIMESTAMP_EASTERN']).dt.round("S") 


# In[10]:


df.to_sql("talkiatry", con ,if_exists='replace', index=False)


# In[11]:


df1 = pd.read_sql(f"""select STATE
                    ,count(distinct APPOINTMENT_ID) as 'Total Bookings'
                    ,sum(MARKETPLACE_BOOKINGS) as 'Marketplace Bookings'
                    ,sum(WIDGET_BOOKINGS) as 'Widget Bookings'
                    from talkiatry group by STATE
                    order by 'Total Bookings','Marketplace Bookings','Widget Bookings' asc;""",con)


# In[12]:


df1


# In[13]:


df1 = df1.sort_values(by=['Total Bookings','Marketplace Bookings','Widget Bookings'], ascending=False)
df1= df1.head(5)


# In[14]:


df1


# In[15]:


df2 = pd.read_sql(f"""select STATE
                    ,count(distinct APPOINTMENT_ID) as 'Total Bookings' 
                    ,sum(MARKETPLACE_BOOKINGS) as 'Marketplace Bookings'
                    ,sum(WIDGET_BOOKINGS) as 'Widget Bookings'
                    from talkiatry group by STATE
                    order by 'Total Bookings','Marketplace Bookings','Widget Bookings' asc;""",con)


# In[16]:


df2 = df2.sort_values(by=['Total Bookings','Marketplace Bookings','Widget Bookings'], ascending=True)
df2= df2.head(5)


# In[17]:


df2


# In[18]:


df3 = pd.read_sql(f"""select STATE, Provider_Full_Name as 'Provider Name'
                    ,count(distinct APPOINTMENT_ID) as 'Total Bookings'
                    ,sum(MARKETPLACE_BOOKINGS) as 'Marketplace Bookings'
                    ,sum(WIDGET_BOOKINGS) as 'Widget Bookings'
                    from talkiatry 
                    group by STATE,Provider_Full_Name;""",con)


# In[19]:


df3


# In[20]:


df3 = df3.sort_values(by=['STATE','Total Bookings'], ascending=[True,False])


# In[21]:


df3


# In[22]:


df4 = pd.read_sql(f"""select SPECIALTY as Specialty, Provider_Full_Name as 'Provider Name' 
                    ,count(distinct APPOINTMENT_ID) as 'Booking Counts Total' 
                    ,sum(MARKETPLACE_BOOKINGS) as 'Marketplace Bookings'
                    ,sum(WIDGET_BOOKINGS) as 'Widget Bookings'
                    from talkiatry 
                    group by SPECIALTY,Provider_Full_Name
                    ;""",con)


# In[23]:


df4 = df4.sort_values(by=['Specialty','Booking Counts Total'], ascending=[True,False])


# In[24]:


df4


# In[25]:


df4 = pd.pivot_table(df4,index=["Specialty","Provider Name"],
               values=['Marketplace Bookings','Booking Counts Total' ,'Widget Bookings'],fill_value=0, margins=True,margins_name='Grand Total',aggfunc=sum)


# In[26]:


df4


# In[27]:


from datetime import datetime, timedelta
day = datetime.today().strftime('%Y-%m-%d')
dt = datetime.strptime(day, '%Y-%m-%d')
end_date = dt - timedelta(days=dt.weekday())
end_date = end_date - timedelta(days=1)
start_date = end_date - timedelta(days=6)
start_date= str(start_date.strftime('%m.%d.%Y'))
end_date= str(end_date.strftime('%m.%d.%Y'))


# In[28]:


date_columns = df.select_dtypes(include=['datetime64[ns, UTC]']).columns
for date_column in date_columns:
    df[date_column] = df[date_column].apply(str)


# In[29]:


df.replace({'NaT': ' '}, inplace=True)


# In[32]:


writer = pd.ExcelWriter(r'C:\Users\Vinayak.Laxmeshwar\Desktop\Talkiatry\Talkiatry Booking Download ' + start_date + ' to ' + end_date+ '.xlsx',mode='w')

# Write each dataframe to a different worksheet.
df1.to_excel(writer, sheet_name='Bookings by Prov by State',index = False,startcol=0, startrow=6)
df2.to_excel(writer, sheet_name='Bookings by Prov by State',index = False,startcol=0, startrow=17)
df3.to_excel(writer, sheet_name = 'Bookings by Prov by State', index = False,startcol=6, startrow=6)
df4.to_excel(writer, sheet_name = 'Bookings by Prov by Spec',startcol=0, startrow=6)
df.to_excel(writer, sheet_name = 'Bookings Download', index = False)



sheet = writer.sheets['Bookings by Prov by State'] 
sheet.write(1,0,"Bookings by Provider by State")

sheet1 = sheet
sheet1.write(2,0,"Week of")

sheet3= sheet
sheet3.write(5,0,'[TOP 5 STATES] Bookings by State')

sheet4 = sheet
sheet4.write(16,0,'[BOTTOM 5 STATES] Bookings by State')

sheet5 = sheet
sheet5.write(5,6,'[ALL STATES] Bookings by Provider by State')



# -------------- for Bookings by Provider by Specialty

sheet6 = writer.sheets['Bookings by Prov by Spec'] 
sheet6.write(1,0,"Bookings by Provider by Specialty")

sheet7 =  writer.sheets['Bookings by Prov by Spec'] 
sheet7.write(2,0,"Week of")

sheet8 = writer.sheets['Bookings by Prov by Spec']
sheet8.write(5,2,'Bookings by Provider by Specialty')

sheet9 = writer.sheets['Bookings by Prov by Spec']
sheet9.write(5,0,'Bookings by Provider by Specialty')

# Close the Pandas Excel writer and output the Excel file.

writer.save()
writer.close()


# In[36]:


import openpyxl
from openpyxl.styles import Font, Color, PatternFill,Alignment, Border, Side



day = datetime.today().strftime('%Y-%m-%d')
dt = datetime.strptime(day, '%Y-%m-%d')
end = dt - timedelta(days=dt.weekday())
start = end - timedelta(days=7)
start =start.strftime('%m/%d/%Y')

#--------Cell Color change
redFill = PatternFill(start_color='79a6d2',
                   end_color='79a6d2',
                   fill_type='solid') #----For "week of"

redFill2 = PatternFill(start_color='4775d1',
                   end_color='4775d1',
                   fill_type='solid')

redFill3 = PatternFill(start_color='eaeffa',
                   end_color='eaeffa',
                   fill_type='solid')

#---------------- Border
sides = Side(border_style=None)
no_border = Border(
    left=sides, 
    right=sides, 
    top=sides, 
    bottom=sides,
)

wb = openpyxl.load_workbook(r'C:\Users\Vinayak.Laxmeshwar\Desktop\Talkiatry\Talkiatry Booking Download ' + start_date + ' to ' + end_date+ '.xlsx')

#-----------------for sheet 1 
ws =wb['Bookings by Prov by State']
ws

#-----------------for sheet 2
wt=wb['Bookings by Prov by Spec']
wt


font_style = Font(name= 'Roboto', size=14)
font_style2 = Font(color='ffffff', bold= True)
align = Alignment(horizontal='center')

A2 = ws['A2']
A2.font = font_style

AB2 = wt['A2']
AB2.font = font_style

A6=ws['A6']
A6.font = font_style2

A17=ws['A17']
A17.font = font_style2

G6=ws['G6']
G6.font = font_style2

C6=wt['A6']
C6.font = font_style2

ws["A6"].alignment = Alignment(horizontal='center')
ws["A17"].alignment = Alignment(horizontal='center')
ws["G6"].alignment = Alignment(horizontal='center')
wt["A6"].alignment = Alignment(horizontal='center')
            

ws['A3'].fill = redFill
wt['A3'].fill = redFill
ws['A6'].fill = redFill2
ws['A17'].fill = redFill2
ws['G6'].fill = redFill2
wt['A6'].fill = redFill2
ws['A7'].fill = redFill3
ws['B7'].fill = redFill3
ws['C7'].fill = redFill3
ws['D7'].fill = redFill3
ws['A18'].fill = redFill3
ws['B18'].fill = redFill3
ws['C18'].fill = redFill3
ws['D18'].fill = redFill3
ws['G7'].fill = redFill3
ws['H7'].fill = redFill3
ws['I7'].fill = redFill3
ws['J7'].fill = redFill3
ws['K7'].fill = redFill3
wt['A7'].fill = redFill3
wt['b7'].fill = redFill3
wt['C7'].fill = redFill3
wt['E7'].fill = redFill3
wt['D7'].fill = redFill3
ws['B3'].fill = redFill
wt['B3'].fill = redFill


ws.merge_cells('A2:AF2')
ws.merge_cells('A6:D6')
ws.merge_cells('A17:D17')
ws.merge_cells('G6:K6')
wt.merge_cells('A6:E6')

ws['B3'].value = start
wt['B3'].value = start


ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 24
ws.column_dimensions['H'].width = 25
ws.column_dimensions['I'].width = 15
ws.column_dimensions['J'].width = 24
ws.column_dimensions['K'].width = 25
wt.column_dimensions['A'].width = 55
wt.column_dimensions['B'].width = 25
wt.column_dimensions['C'].width = 20
wt.column_dimensions['D'].width = 25
wt.column_dimensions['E'].width = 25

for row in wt:
        for cell in row:
            # Apply colorless and borderless styles
            cell.border = no_border
            
for row in ws:
        for cell in row:
            # Apply colorless and borderless styles
            cell.border = no_border
            
for row in wt.iter_rows(min_row=wt.max_row,max_row=wt.max_row, min_col=1,max_col =wt.max_column):
    for cell in row:    
        cell.fill = redFill3 
        
for row in wt.iter_rows(min_row=wt.max_row,max_row=wt.max_row, min_col=1,max_col =wt.max_column):
    for cell in row:    
        cell.font = Font(bold= True)
    
            


wb.save(r'C:\Users\Vinayak.Laxmeshwar\Desktop\Talkiatry\Talkiatry Booking Download ' + start_date + ' to ' + end_date+ '.xlsx')


# In[ ]:


import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os.path

email = 'vinayak.laxmeshwar@zocdoc.com'
password = 'rubzzqumlkcrzagk'
to = "rachel.lomax@zocdoc.com,,lillian.pierce@zocdoc.com"
cc = "tanvi.malik@zocdoc.com,vinayak.laxmeshwar@zocdoc.com,enterprisesupport@zocdoc.com"
subject = 'Talkiatry Bookings Download- ' + start_date + ' to ' + end_date+'.'
message = """Hi All,

Please find attached the bookings download data for Talkiatry.

Feel free to reach out if you have any questions.


Thanks & Regards,
Vinayak R Laxmeshwar"""

file_location = r'C:\Users\Vinayak.Laxmeshwar\Desktop\Talkiatry\Talkiatry Booking Download ' + start_date + ' to ' + end_date+ '.xlsx'

rcpt = cc.split(",") + [to]

msg = MIMEMultipart()
msg['From'] = email
msg['To'] = to
msg['Cc'] = cc
msg['Subject'] = subject

msg.attach(MIMEText(message, 'plain'))

# Setup the attachment
filename = os.path.basename(file_location)
attachment = open(file_location, "rb")
part = MIMEBase('application', 'octet-stream')
part.set_payload(attachment.read())
encoders.encode_base64(part)
part.add_header('Content-Disposition', "attachment; filename= %s" % filename)

# Attach the attachment to the MIMEMultipart object
msg.attach(part)

server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login(email, password)
text = msg.as_string()
server.sendmail(email, rcpt, text)
server.quit()


# In[ ]:




