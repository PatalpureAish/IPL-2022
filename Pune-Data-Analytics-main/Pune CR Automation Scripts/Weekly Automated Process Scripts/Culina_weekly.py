#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
from pandas import DataFrame
import snowflake.connector
import sqlite3
from datetime import datetime, timedelta
import openpyxl
datetime, timedelta
from openpyxl.styles import Font, Color, PatternFill,Alignment 
import numpy as np

# Gets the version
conn = snowflake.connector.connect(
    user="vinayak.laxmeshwar@zocdoc.com",
    password="do not enter your password. it will authenticate with the browser, however this parameter cannot be an empty string.",
    account="zocdoc_001.us-east-1",
    authenticator="externalbrowser",
    database="cistern"
    )
con = sqlite3.connect('rhcpd5.db')


# In[5]:


day = datetime.today().strftime('%Y-%m-%d')
dt = datetime.strptime(day, '%Y-%m-%d')
end = dt - timedelta(days=dt.weekday()+1)
print(end.strftime('%Y-%m-%d'))


# In[3]:


df = pd.read_sql(f"""SELECT
    p.first_name
    ,p.last_name
    ,p.monolith_professional_id
    ,p.npi
    ,count(a.location_id) AS Locations
    ,sum(a.DAY_7_NUM_HOURS_WITH_AVAILABILITY) AS one_week_availability
    ,sum(a.DAY_28_NUM_HOURS_WITH_AVAILABILITY) AS four_week_availability

FROM AVAILABILITY.PROVIDER_LOCATION_UPPER_FUNNEL_AVAILABILITY_BY_DAY AS a 

LEFT JOIN provider.provider_vw as p
ON a.provider_id = p.provider_id


WHERE p.monolith_provider_id = '65520'
AND a.date_utc = '{end.strftime('%Y-%m-%d')}'
AND p.status = 'Approved Application' 

GROUP BY 1,2,3,4

ORDER BY 1,2""",conn)


# In[4]:


df


# In[5]:


day = datetime.today().strftime('%Y-%m-%d')
dt = datetime.strptime(day, '%Y-%m-%d')
end = dt - timedelta(days=dt.weekday())
end_date = str(end.strftime('%Y.%m.%d'))
end_date2 = str(end.strftime('%Y/%m/%d'))


# In[6]:


end_date2


# In[7]:


writer = pd.ExcelWriter(r'C:\Users\Vinayak.Laxmeshwar\Desktop\Culina\Culina Health Provider Availability_' + end_date+ '.xlsx',mode='w')

# Write each dataframe to a different worksheet.
df.to_excel(writer, sheet_name='Availability',index = False,startcol=1, startrow=3)

sheet = writer.sheets['Availability'] 
sheet.write(1,1,"Culina Health Hours of Availability per Provider")

sheet = writer.sheets['Availability'] 
sheet.write(2,1,"Data As Of " +end_date2+ ".")



writer.save()
writer.close()


# In[8]:


import openpyxl
from openpyxl.styles import Font, Color, PatternFill,Alignment, Border, Side
wb = openpyxl.load_workbook(r'C:\Users\Vinayak.Laxmeshwar\Desktop\Culina\Culina Health Provider Availability_' + end_date+ '.xlsx')


ws =wb['Availability']
ws

font_style = Font(bold= True)
font_style2 = Font(italic=True)

B2 = ws['B2']
B2.font = font_style

B3 = ws['B3']
B3.font = font_style2


ws.merge_cells('B2:H2')
ws.merge_cells('B3:H3')




wb.save(r'C:\Users\Vinayak.Laxmeshwar\Desktop\Culina\Culina Health Provider Availability_' + end_date+ '.xlsx')


# In[10]:


import openpyxl
from openpyxl.styles import Font, Color, PatternFill,Alignment, Border, Side
wb = openpyxl.load_workbook(r'C:\Users\Vinayak.Laxmeshwar\Desktop\Culina\Culina Health Provider Availability_' + end_date+ '.xlsx')


redFill = PatternFill(start_color='BDD7EE',
                   end_color='BDD7EE',
                   fill_type='solid') 

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

ws =wb['Availability']
ws

ws['B4'].fill = redFill
ws['C4'].fill = redFill
ws['D4'].fill = redFill
ws['E4'].fill = redFill
ws['F4'].fill = redFill
ws['G4'].fill = redFill
ws['H4'].fill = redFill

ws["B2"].alignment = Alignment(horizontal='center')
ws["B3"].alignment = Alignment(horizontal='center')

ws.column_dimensions['B'].width = 11
ws.column_dimensions['C'].width = 11
ws.column_dimensions['D'].width = 17
ws.column_dimensions['E'].width = 11
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 28
ws.column_dimensions['H'].width = 28


ws['B4']='First Name'
ws['C4']='Last Name'
ws['D4']='Monolith Zocdoc ID'
ws['F4']='Location Count'
ws['G4']='One Week Availability Hours'
ws['H4']='Four Week Availability Hours'

for row in ws.iter_rows(min_row = 4 ,max_row = ws.max_row,min_col = 2,max_col  = ws.max_column):
        for cell in row:
            # Apply colorless and borderless styles
            cell.border = thin_border

wb.save(r'C:\Users\Vinayak.Laxmeshwar\Desktop\Culina\Culina Health Provider Availability_' + end_date+ '.xlsx')


# In[49]:


import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os.path

email = 'vinayak.laxmeshwar@zocdoc.com'
password = 'rubzzqumlkcrzagk'
to = "jubrine.alva@zocdoc.com"
cc = "morrisa.cohen@zocdoc.com,vinayak.laxmeshwar@zocdoc.com,tanvi.malik@zocdoc.com,partner-culinahealth@zocdoc.com"
subject = 'Culina Health Availability by Provider '+ end_date+'.'
message = """Hi All,

Please find attached the Culina Health Availability by Provider file.

Feel free to reach out if you have any questions.


Thanks & Regards,
Vinayak R Laxmeshwar"""

file_location = r'C:\Users\Vinayak.Laxmeshwar\Desktop\Culina\Culina Health Provider Availability_' + end_date+ '.xlsx'

rcpt = cc.split(",") + [to]

msg = MIMEMultipart()
msg['From'] = email
msg['To'] = to
msg['Cc'] = cc
msg['Subject'] = subject

msg.attach(MIMEText(message, 'plain'))

# Setup the attachment
filename = os.path.basename(file_location)
attachment = open(file_location, "rb")
part = MIMEBase('application', 'octet-stream')
part.set_payload(attachment.read())
encoders.encode_base64(part)
part.add_header('Content-Disposition', "attachment; filename= %s" % filename)

# Attach the attachment to the MIMEMultipart object
msg.attach(part)

server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login(email, password)
text = msg.as_string()
server.sendmail(email, rcpt, text)
server.quit()


# In[ ]:




