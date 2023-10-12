#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
from pandas import DataFrame
import snowflake.connector
##import sqlite3
from datetime import datetime, timedelta
import openpyxl
datetime, timedelta
from openpyxl.styles import Font, Color, PatternFill,Alignment 
import numpy as np

# Gets the version
conn = snowflake.connector.connect(
    user="vinayak.laxmeshwar@zocdoc.com",
    password="do not enter your password. it will authenticate with the browser, however this parameter cannot be an empty string.",
    account="zocdoc_001.us-east-1",
    authenticator="externalbrowser",
    database="cistern"
    )


# In[2]:


from datetime import datetime, date, timedelta

# Get the current date
current_date = date.today()

# Calculate the first date of the prior month
if current_date.month == 1:
    first_date = date(current_date.year - 1, 12, 1)
else:
    first_date = date(current_date.year, current_date.month - 1, 1)

# Calculate the last date of the prior month
last_date = current_date.replace(day=1) 

# Print the first and last dates
print("First Date:", first_date)
print("Last Date:", last_date)


# In[3]:


df=pd.read_sql(f"""SELECT
    concat(a.provider_first_name,' ',a.provider_last_name) as Provider_Name
    ,a.practice_name as Practice_Name
    ,to_char (date_trunc('day',a.appointment_created_timestamp_local) ,'MM-DD-YYYY') as Appointment_Created_Date_Local
    ,to_char (date_trunc('day',a.latest_appointment_time_local) ,'MM-DD-YYYY') as Latest_Appointment_Date_Local
    ,CASE
        WHEN a.is_non_chargeable_cancellation_from_pe_vw = TRUE then 'Non_Chargable_Patient_Cancellation'
        WHEN a.cancellation_reason like 'Provider_ReschedulingPatient' then 'Rescheduled_by_Provider'
        WHEN a.cancellation_reason like 'Patient_%' then 'Cancelled_by_Patient'
 		WHEN a.cancellation_reason like 'Provider_%' then 'Cancelled_by_Provider'
        WHEN a.appointment_outcome = 'RealizedAppointment' then 'Confirmed'
        WHEN a.appointment_outcome = 'BookingFailed' then 'Rescheduling Error'
        WHEN a.appointment_outcome is null then 'Upcoming Appointment'
        else a.appointment_outcome end as Booking_Outcome
    ,count(a.APPOINTMENT_ID) as Appointment_Count

FROM appointment.appointment_summary_commercial_vw as a

WHERE a.monolith_parent_strategic_id = 375
AND is_created_appointment
AND is_premium_booking = TRUE
AND is_non_chargeable_cancellation_from_pe_vw = FALSE
AND a.procedure_id not in ('pc_3BvV0dlIbkC_Hj_HQ8V4nB', 'pc_m459Tgq50kGv-faPIAb_3h', 'pc_wloLfj5vbUmMiwvX709pXx', 'pc_m0Wv6TPsc0iu-5Ju3vj7wh')
AND date_trunc('month', a.APPOINTMENT_CREATED_TIMESTAMP_LOCAL) = dateadd('month', -1, date_trunc('month', current_date()))

GROUP BY 1,2,3,4,5
ORDER BY 1,2,3,4
""",conn)


# In[4]:


df


# In[5]:


import time
ym = first_date.strftime("%B %Y")


# In[6]:


ym


# In[7]:


import xlwings as xw

wb = xw.Book(r'C:\Users\Vinayak.Laxmeshwar\Desktop\BCHP Monthly\BCHP_Invoice Details_Template.xlsx')
xw.sheets['BCHP Invoice Details'].range('B10').options(index=False).value = df

wb.save(r'C:\Users\Vinayak.Laxmeshwar\Desktop\BCHP Monthly\BCHP_Invoice Details- ' +ym +'.xlsx')
wb.close()


# In[8]:


from datetime import datetime, timedelta
input_dt = datetime.today()
res = input_dt.replace(day=1)
res  = res - timedelta(days=1)
res = res.replace(day=1)
res = res.strftime("%B %d, %Y")


# In[9]:


res


# In[10]:


input_dt = datetime.today()
last = input_dt.replace(day=1)
last = last - timedelta(days=1)
last = last.strftime("%B %d, %Y")


# In[11]:


last


# In[12]:


wb = openpyxl.load_workbook(r'C:\Users\Vinayak.Laxmeshwar\Desktop\BCHP Monthly\BCHP_Invoice Details- ' +ym +'.xlsx')
ws = wb['BCHP Invoice Details']
ws['C5'] =  res + ' - ' + last



wb.save(r'C:\Users\Vinayak.Laxmeshwar\Desktop\BCHP Monthly\BCHP_Invoice Details- ' +ym +'.xlsx')


# In[13]:


from openpyxl.styles import Border,Side
wb = openpyxl.load_workbook(r'C:\Users\Vinayak.Laxmeshwar\Desktop\BCHP Monthly\BCHP_Invoice Details- ' +ym +'.xlsx')
ws = wb['BCHP Invoice Details']

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for row in ws.iter_rows(min_row = 10 ,max_row = ws.max_row,min_col = 2,max_col  = 7):
        for cell in row:
            # Apply colorless and borderless styles
            cell.border = thin_border
            
wb.save(r'C:\Users\Vinayak.Laxmeshwar\Desktop\BCHP Monthly\BCHP_Invoice Details- ' +ym +'.xlsx')


# In[ ]:




