#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
from pandas import DataFrame
import snowflake.connector
##import sqlite3
from datetime import datetime, timedelta
import openpyxl
datetime, timedelta
from openpyxl.styles import Font, Color, PatternFill,Alignment,Border 
import numpy as np

# Gets the version
conn = snowflake.connector.connect(
    user="vinayak.laxmeshwar@zocdoc.com",
    password="do not enter your password. it will authenticate with the browser, however this parameter cannot be an empty string.",
    account="zocdoc_001.us-east-1",
    authenticator="externalbrowser",
    database="cistern"
    )


# In[2]:


from datetime import datetime, date, timedelta

# Get the current date
current_date = date.today()

# Calculate the first date of the prior month
if current_date.month == 1:
    first_date = date(current_date.year -1, 12, 1)
else:
    first_date = date(current_date.year, current_date.month - 1, 1)

# Calculate the last date of the prior month
last_date = current_date.replace(day=1) 

# Print the first and last dates
print("First Date:", first_date)
print("Last Date:", last_date)


# In[3]:


df=pd.read_sql(f"""
           with appt_temp_table as(
            select SUPERVISING_PROVIDER_ID
                 , PROVIDER_ID
                 , LOCATION_ID
                 , IS_NEW_TO_PROVIDER_FROM_PE_VW
                 ,is_premium_booking
                 , APPOINTMENT_ID
                 , APPOINTMENT_OUTCOME
                 , LATEST_APPOINTMENT_TIME_LOCAL
                 , APPOINTMENT_CREATED_TIMESTAMP_LOCAL
            from APPOINTMENT.APPOINTMENT_SUMMARY_COMMERCIAL_VW a
                     LEFT JOIN provider_analytics.practice_parent_entity_mapping_vw as e
                               ON e.practice_id = a.practice_id
            where IS_CREATED_APPOINTMENT
              and a.procedure_id not in
                  ('pc_3BvV0dlIbkC_Hj_HQ8V4nB', 'pc_m459Tgq50kGv-faPIAb_3h', 'pc_wloLfj5vbUmMiwvX709pXx',
                   'pc_m0Wv6TPsc0iu-5Ju3vj7wh')
              and a.is_premium_booking = TRUE
              and e.SALESFORCE_ULTIMATE_PARENT_ACCOUNT_ID = '0016000000JjDaSAAV'
              and (date_trunc('month', a.LATEST_APPOINTMENT_TIME_LOCAL) = '{first_date.strftime('%Y-%m-%d')}'
                or
                   date_trunc('month', a.APPOINTMENT_CREATED_TIMESTAMP_LOCAL) = '{first_date.strftime('%Y-%m-%d')}'
                ))

            ,appointments as
                (select SUPERVISING_PROVIDER_ID           as provider_id
                      , a.PROVIDER_ID                     as resource_id
                      , a.LOCATION_ID
                      , count(*)                          as total_premium_appointments
                      , count(distinct case
                                           when a.appointment_outcome = 'RealizedAppointment' then APPOINTMENT_ID
                                           else null end) as realized_appointments
                 from appt_temp_table a

                 where date_trunc('month', a.LATEST_APPOINTMENT_TIME_LOCAL) = '{first_date.strftime('%Y-%m-%d')}'
                 group by 1, 2, 3)
               , bookings as
                -- Simple bookings data, set to the last full month
                (select b.SUPERVISING_PROVIDER_ID         as provider_id
                      , b.PROVIDER_ID                     as resource_id
                      , b.LOCATION_ID
                      , count(*)                          as total_premium_bookings

                 from appt_temp_table b
                 where date_trunc('month', b.APPOINTMENT_CREATED_TIMESTAMP_LOCAL) = '{first_date.strftime('%Y-%m-%d')}'
                 group by 1, 2, 3)

               , provider_location_base as
                (select a.*
                      , bookings.total_premium_bookings
                      , appointments.total_premium_appointments
                      , appointments.realized_appointments


                 from (select distinct PROVIDER_ID, location_id, resource_id
                       from appointments
                       union
                       select distinct PROVIDER_ID, location_id, resource_id
                       from bookings) a
                          left join appointments using (PROVIDER_ID, resource_id, location_id)
                          left join bookings using (PROVIDER_ID, resource_id, location_id))

            --    select * from provider_location_base;

               , resource_identification as
                -- this cte identifies if a provider is also a resource, so they can be removed from our main providers
                (select distinct RESOURCE_ID
                 from provider.provider_resource_mapping prm
                          left join PROVIDER.PROVIDER provider_map ON provider_map.provider_id = prm.PROVIDER_ID

                          LEFT JOIN provider_analytics.practice_parent_entity_mapping_vw as e
                                    ON e.practice_id = provider_map.practice_id
                 where e.SALESFORCE_ULTIMATE_PARENT_ACCOUNT_ID = '0016000000JjDaSAAV')

               , provider_base as
                -- This gets our base of providers and we are joining provider data in order to remove churned
                -- We need to use distinct here due to the Medical Resource_id field
                (select distinct prm.PROVIDER_ID
                               , r.RESOURCE_ID
                               , l.LOCATION_ID
                 from provider.provider_resource_mapping prm
                          left join PROVIDER.PROVIDER provider_map ON provider_map.provider_id = prm.PROVIDER_ID
                          left join PROVIDER.PROVIDER resource_map on resource_map.PROVIDER_ID = prm.resource_id

                          left join provider.LOCATION l on provider_map.PRACTICE_ID = l.PRACTICE_ID

                          LEFT JOIN provider_analytics.practice_parent_entity_mapping_vw as e
                                    ON e.practice_id = provider_map.practice_id

                     -- joining the identified resources here to remove them from the provider side
                     -- this join is just for the where clause
                          left join resource_identification r on r.RESOURCE_ID = prm.PROVIDER_ID

                 where e.SALESFORCE_ULTIMATE_PARENT_ACCOUNT_ID = '0016000000JjDaSAAV'
                   and provider_map.STATUS_ID = 2
                   -- Removing churned resourced
                   and (resource_map.STATUS_ID = 2 or resource_map.STATUS_ID is null)

                   -- remove resource
                   and r.RESOURCE_ID is null
                   and l.DELETED_TIME_UTC is null)

            --    select * from provider_base;

               , missing_providers as
                -- Finds the gaps where providers many not have bookings to union in next CTE
                (select a.PROVIDER_ID
                      , null as LOCATION_ID
                      , a.resource_id
                      , 0    as total_premium_bookings
                      , 0    as total_premium_appointments
                      , 0    as realized_appointments
                 from provider_base a
                 where (a.PROVIDER_ID, coalesce(resource_id, a.PROVIDER_ID))
                           not in (select distinct PROVIDER_ID, resource_id from provider_location_base))

            --    select * from provider_location_base;

               , unioned_provider_map as
                (select provider_id,
                        resource_id,
                        LOCATION_ID,
                        total_premium_bookings,
                        total_premium_appointments,
                        realized_appointments
                 from provider_location_base
                 union
                 select provider_id,
                        resource_id,
                        LOCATION_ID,
                        total_premium_bookings,
                        total_premium_appointments,
                        realized_appointments
                 from missing_providers)

            -- select *
            -- from unioned_provider_map;


               , resource_mapping_with_metadata as
                -- Bringing it together with metadata
                (select distinct concat(provider_map.first_name, ' ', provider_map.last_name)    as provider_name
                               , provider_map.MONOLITH_PROFESSIONAL_ID                           as professional_id
                               , provider_map.TITLE                                              as provider_type
                               , e.MONOLITH_PROVIDER_ID                                          as practice_id
                               , location_id
                               , resource_map.MONOLITH_PROFESSIONAL_ID                           as resource_id
                               , concat(resource_map.first_name, ' ', resource_map.last_name)    as resource_name
                               , resource_map.TITLE                                              as resource_type
                               , total_premium_bookings
                               , total_premium_appointments
                               , realized_appointments
                               , count(distinct coalesce(resource_name, provider_name))
                                       over (partition by provider_map.MONOLITH_PROFESSIONAL_ID) as resource_count
                 from unioned_provider_map upm
                          left join PROVIDER.PROVIDER provider_map ON provider_map.provider_id = upm.provider_id
                          left join PROVIDER.PROVIDER resource_map on resource_map.PROVIDER_ID = upm.resource_id
                          LEFT JOIN provider_analytics.practice_parent_entity_mapping_vw as e
                                    ON e.practice_id = provider_map.practice_id)
               , final as
                -- This last cte just removes providers that are mapped as their own resource due to provider_id = supervising_provider_id
                (select a.provider_name
                      , professional_id
                      , provider_type
                      , a.practice_id
                      , case when l.name is null then l.address1 else l.name end                 as location_name
                      , case when resource_count > 1 then True else false end                    as has_mapped_resources
                      , case when resource_id = professional_id then null else resource_id end   as resource_id
                      , case when resource_id = professional_id then null else resource_name end as resource_name
                      , case when resource_id = professional_id then null else resource_type end as resource_type
                      , total_premium_bookings
                      , total_premium_appointments
                      --calculate realization rate without upcoming appointments, null for providers with zero bookings
                      , div0(realized_appointments, nullifzero(total_premium_appointments))              as realization_rate
                 from resource_mapping_with_metadata a
                          left join PROVIDER.LOCATION l
                                    on a.LOCATION_ID = l.LOCATION_ID)
            select *
            from final
            order by provider_name, resource_name, location_name
            ;""",conn)


# In[4]:


df


# In[5]:


import time
ym = first_date.strftime("%B %Y")


# In[6]:


ym


# In[7]:


import xlwings as xw

wb = xw.Book(r'C:\Users\Vinayak.Laxmeshwar\Desktop\TDS_Monthly\TDS Resource Mapping Template.xlsx')
xw.sheets['TDS - Resource Mapping Summary'].range('A11').options(index=False).value = df

wb.save(r'C:\Users\Vinayak.Laxmeshwar\Desktop\TDS_Monthly\TDS Double Mapped Resources Report- ' +ym +'.xlsx')
wb.close()


# In[8]:


from datetime import datetime, timedelta
input_dt = datetime.today()
res = input_dt.replace(day=1)
res  = res - timedelta(days=1)
res = res.replace(day=1)
res = res.strftime("%B %d, %Y")


# In[9]:


res


# In[10]:


input_dt = datetime.today()
last = input_dt.replace(day=1)
last = last - timedelta(days=1)
last = last.strftime("%B %d, %Y")


# In[11]:


wb = openpyxl.load_workbook(r'C:\Users\Vinayak.Laxmeshwar\Desktop\TDS_Monthly\TDS Double Mapped Resources Report- ' +ym +'.xlsx')
ws = wb['TDS - Resource Mapping Summary']
ws['C6'] =  res + ' - ' + last



wb.save(r'C:\Users\Vinayak.Laxmeshwar\Desktop\TDS_Monthly\TDS Double Mapped Resources Report- ' +ym +'.xlsx')


# In[12]:


from openpyxl.styles import Side
wb = openpyxl.load_workbook(r'C:\Users\Vinayak.Laxmeshwar\Desktop\TDS_Monthly\TDS Double Mapped Resources Report- ' +ym +'.xlsx')
ws = wb['TDS - Resource Mapping Summary']

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for row in ws.iter_rows(min_row = 10 ,max_row = ws.max_row,min_col = 1,max_col  = 13):
        for cell in row:
            # Apply colorless and borderless styles
            cell.border = thin_border
            
wb.save(r'C:\Users\Vinayak.Laxmeshwar\Desktop\TDS_Monthly\TDS Double Mapped Resources Report- ' +ym +'.xlsx')


# In[ ]:





# In[ ]:




