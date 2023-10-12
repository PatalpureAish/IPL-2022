#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
from pandas import DataFrame
import snowflake.connector
import sqlite3
from datetime import datetime, timedelta
import openpyxl
datetime, timedelta
from openpyxl.styles import Font, Color, PatternFill,Alignment,Border 
import numpy as np

# Gets the version
conn = snowflake.connector.connect(
    user="vinayak.laxmeshwar@zocdoc.com",
    password="do not enter your password. it will authenticate with the browser, however this parameter cannot be an empty string.",
    account="zocdoc_001.us-east-1",
    authenticator="externalbrowser",
    database="cistern"
    )

con = sqlite3.connect('rhcpd6.db')


# In[2]:


from datetime import datetime, date, timedelta

# Get the current date
current_date = date.today()

# Calculate the first date of the prior month
if current_date.month == 1:
    first_date = date(current_date.year -1, 12, 1)
else:
    first_date = date(current_date.year, current_date.month - 1, 1)

# Calculate the last date of the prior month
last_date = current_date.replace(day=1) 

# Print the first and last dates
print("First Date:", first_date)
print("Last Date:", last_date)


# In[3]:


df=pd.read_sql(f"""
                SELECT 
                    DISTINCT ps.review_id
                    ,pr.PRACTICE_NAME as practice_name
                    ,pa.MONOLITH_PROFESSIONAL_ID as provider_id
                    ,pa.PROVIDER_FIRST_NAME AS providerfirstname
                    ,pa.PROVIDER_LAST_NAME AS providerlastname
                    ,TO_CHAR(pa.APPOINTMENT_CREATED_TIMESTAMP_UTC, 'YYYY-MM-DD HH24:MI:SS') AS bookingtime
                    ,TO_CHAR(pa.APPOINTMENT_OUTCOME_TIMESTAMP_UTC, 'YYYY-MM-DD HH24:MI:SS') AS finalappointmenttime
                    ,TO_CHAR(ps.creation_timestamp_utc, 'YYYY-MM-DD HH24:MI:SS') AS reviewdate
                    ,CASE
                        WHEN pa.is_non_chargeable_cancellation_from_pe_vw = TRUE then 'Non_Chargable_Patient_Cancellation'
                        WHEN pa.cancellation_reason like 'Patient_%' then 'Cancelled_by_Patient'
                        WHEN pa.cancellation_reason like 'Provider_%' then 'Cancelled_by_Provider'
                        WHEN pa.appointment_outcome = 'RealizedAppointment' then 'Confirmed'
                        WHEN pa.appointment_outcome = 'BookingFailed' then 'Rescheduling Error'
                        WHEN pa.appointment_outcome is null then 'Upcoming Appointment'
                        else pa.appointment_outcome end as Booking_Outcome
                    ,ps.OVERALL_RATING AS overall
                    ,ps.WAITTIME_RATING
                    ,ps.BEDSIDE_RATING AS bedsidemanner
                    ,ps.COMMENT
                    ,CASE
                        WHEN ps.IS_PARTNER_REVIEW = TRUE then 'Third Party Review'
                        WHEN ps.IS_PARTNER_REVIEW = FALSE then 'Zocdoc Review'
                        ELSE NULL END AS Review_Type

                FROM patient.review_latest_vw AS ps

                LEFT JOIN appointment.appointment_summary_commercial_vw pa
                    ON pa.appointment_id = ps.appointment_id

                left join PROVIDER.PROVIDER AS pr
                    ON pa.monolith_professional_id = pr.monolith_professional_id

                LEFT JOIN provider.strategic_info_vw as s
                 on pr.strategic_id = s.monolith_strategic_id

                JOIN provider_analytics.practice_parent_entity_mapping_vw as e
                ON e.practice_id=pr.practice_id

                --Put '--' in front of both lines below for LPG. For Health System parent accounts, use the first field, for Health System child accounts, use the second field
                    --WHERE s.monolith_parent_strategic_id = 250
                    --WHERE s.monolith_strategic_id = XXX

                --Put '--' in front of both lines below for Health System. For LPGs, use one of the following fields
                    --WHERE pr.MONOLITH_PROVIDER_ID IN ('344253') -- Change Practice ID here
                    --WHERE pr.monolith_professional_id = 344253
                    WHERE e.SALESFORCE_ULTIMATE_PARENT_ACCOUNT_ID like '%0010d00001TSV0PAAX%'

                --live doctors only, put '--' in front of below line for all providers
                    --AND pr.status_id IN (2, 60)

                AND ps.OVERALL_RATING IS NOT NULL


                --Adjust time frame here--
                AND ps.creation_timestamp_utc BETWEEN '{first_date.strftime('%Y-%m-%d')}' and '{last_date.strftime('%Y-%m-%d')}'

                ORDER BY bookingtime ASC""",conn)


# In[4]:


df


# In[5]:


df.to_sql("Primecare", con ,if_exists='replace', index=False)


# In[6]:


df = pd.read_sql(f"""select PRACTICE_NAME, PROVIDER_ID,PROVIDERFIRSTNAME,PROVIDERLASTNAME,BOOKINGTIME,FINALAPPOINTMENTTIME,REVIEWDATE,BOOKING_OUTCOME,OVERALL,WAITTIME_RATING,BEDSIDEMANNER,COMMENT,REVIEW_TYPE
 from Primecare""",con)


# In[7]:


import time
ym = first_date.strftime("%B %Y")


# In[8]:


import xlwings as xw

wb = xw.Book(r'C:\Users\Vinayak.Laxmeshwar\Desktop\Primecare\PrimeCare Reviews_Jan 2023_no comments.xlsx')
xw.sheets['PrimeCare Reviews'].range('B2').options(index=False).value = df

wb.save(r'C:\Users\Vinayak.Laxmeshwar\Desktop\Primecare\PrimeCare Reviews - ' +ym +'.xlsx')
wb.close()


# In[ ]:


from openpyxl.styles import Side
wb = openpyxl.load_workbook(r'C:\Users\Vinayak.Laxmeshwar\Desktop\Primecare\PrimeCare Reviews - ' +ym +'.xlsx')
ws = wb['PrimeCare Reviews']

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for row in ws.iter_rows(min_row = 2 ,max_row = ws.max_row,min_col = 2,max_col  = 13):
        for cell in row:
            # Apply colorless and borderless styles
            cell.border = thin_border
            
wb.save(r'C:\Users\Vinayak.Laxmeshwar\Desktop\Primecare\PrimeCare Reviews - ' +ym +'.xlsx')


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




